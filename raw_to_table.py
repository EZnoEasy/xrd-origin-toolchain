"""
Bruker RAW V3/V4 --> 2theta & Relative Intensity Table
拖拽 .raw 到 .bat 上即可，自动在同目录生成 {文件名}_table.csv 和 {文件名}_table.xlsx
"""

import struct
import sys
import os

try:
    import numpy as np
except ImportError:
    print("ERROR: numpy required. Run: pip install numpy")
    input("Press Enter to exit...")
    sys.exit(1)

# ── Binary read helpers ──────────────────────────────────────
def _r32(data, off):
    return struct.unpack('<I', data[off:off+4])[0]

def _f64(data, off):
    return struct.unpack('<d', data[off:off+8])[0]

def _f32(data, off):
    return struct.unpack('<f', data[off:off+4])[0]

def _str(data, off, n):
    return data[off:off+n].decode('latin-1', errors='replace').rstrip('\x00').strip()


def parse_raw(filepath):
    """Parse Bruker RAW1.01 --> (angles_2theta, intensities, meta)"""
    with open(filepath, 'rb') as f:
        data = f.read()

    R = 712  # Range header offset (verified for V3/V4)

    measure_date = _str(data, 16, 10)
    measure_time = _str(data, 26, 10)
    sample_id    = _str(data, 326, 60)
    comment      = _str(data, 386, 160)

    hdr_len   = _r32(data, R)
    steps     = _r32(data, R + 4)
    start_2theta = _f64(data, R + 16)

    # Step size -- multi-strategy
    step_size = 0.01637
    try:
        v = _f64(data, R + 176)
        if 0.001 < v < 0.2:
            step_size = v
    except:
        pass
    if step_size == 0.01637:
        try:
            v = _f64(data, 888)
            if 0.001 < v < 0.2:
                step_size = v
        except:
            pass

    high_voltage  = _f32(data, R + 96)
    time_per_step = _f32(data, R + 192)
    used_lambda   = _f64(data, R + 240)

    # Data region: from end of file backward
    total_len = len(data)
    DATA_START = total_len - steps * 4

    try:
        first_val = struct.unpack('<f', data[DATA_START:DATA_START+4])[0]
    except:
        first_val = 0

    if not (50 < first_val < 200000):
        for off in range(total_len - steps * 4,
                         min(total_len - steps * 4 + 400, total_len - 4), 1):
            try:
                v = struct.unpack('<f', data[off:off+4])[0]
                if 50 < v < 200000:
                    DATA_START = off
                    break
            except:
                continue

    intensities = np.array(
        struct.unpack(f'<{steps}f', data[DATA_START:DATA_START + steps * 4]),
        dtype=np.float64
    )

    if steps > 1 and step_size > 0:
        angles = np.linspace(start_2theta,
                              start_2theta + step_size * (steps - 1),
                              steps)
    else:
        angles = np.array([start_2theta])

    end_2theta = start_2theta + step_size * (steps - 1) if steps > 1 else start_2theta

    meta = {
        'filename': os.path.splitext(os.path.basename(filepath))[0],
        'num_points': steps,
        'start': start_2theta,
        'end': end_2theta,
        'step': step_size,
        'wavelength': used_lambda if 0.5 < used_lambda < 3.0 else 1.5406,
        'voltage': high_voltage,
        'date': measure_date,
        'time': measure_time,
        'sample_id': sample_id,
        'comment': comment,
    }
    return angles, intensities, meta


def main():
    if len(sys.argv) < 2:
        print("=" * 60)
        print("  Bruker RAW --> 2theta + Relative Intensity Table")
        print("=" * 60)
        print()
        print("  用法: 将 .raw 文件拖到 拖入RAW出表格.bat 上即可")
        print()
        input("Press Enter to exit...")
        sys.exit(0)

    raw_path = sys.argv[1].strip('"')
    if not os.path.exists(raw_path):
        print(f"ERROR: File not found: {raw_path}")
        input("Press Enter to exit...")
        sys.exit(1)

    raw_path = os.path.abspath(raw_path)
    out_dir  = os.path.dirname(raw_path)

    if not raw_path.lower().endswith('.raw'):
        print(f"WARNING: File does not have .raw extension: {raw_path}")
        print("Continuing anyway...")
        print()

    print("=" * 60)
    print(f"  Parsing: {os.path.basename(raw_path)}")
    print("=" * 60)
    print()

    try:
        angles, intensities, meta = parse_raw(raw_path)
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()
        input("Press Enter to exit...")
        sys.exit(1)

    # ── Relative intensity (%) ────────────────────────────────
    imax = intensities.max()
    rel_intensity = intensities / imax * 100.0 if imax > 0 else intensities

    # ── Metadata ──────────────────────────────────────────────
    print(f"  Sample:      {meta['sample_id']}")
    print(f"  Date:        {meta['date']} {meta['time']}")
    print(f"  Range:       {meta['start']:.4f} ~ {meta['end']:.4f} deg (2theta)")
    print(f"  Step:        {meta['step']:.6f} deg")
    print(f"  Points:      {meta['num_points']}")
    print(f"  Wavelength:  {meta['wavelength']:.4f} A")
    print(f"  Voltage:     {meta['voltage']:.0f} kV")
    if meta['comment']:
        print(f"  Comment:     {meta['comment']}")
    print()

    # ── Table preview (first 10 rows) ─────────────────────────
    print(f"  {'#':>5}  {'2theta(deg)':>12}  {'Rel.Int(%)':>12}")
    print(f"  {'-'*5}  {'-'*12}  {'-'*12}")
    preview = min(10, len(angles))
    for idx in range(preview):
        print(f"  {idx+1:>5}  {angles[idx]:>12.4f}  {rel_intensity[idx]:>12.2f}")
    if len(angles) > 10:
        print(f"  {'...':>5}  {'...':>12}  {'...':>12}")
        print(f"  ({len(angles)} rows total, first 10 shown above)")
    print()

    # ── Output files ──────────────────────────────────────────
    base = meta['filename']

    # CSV
    csv_path = os.path.join(out_dir, f"{base}_table.csv")
    with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
        f.write("2theta_deg,Rel_Intensity_pct\n")
        for a, r in zip(angles, rel_intensity):
            f.write(f"{a:.4f},{r:.2f}\n")
    print(f"  [OK] CSV: {csv_path}")

    # XLSX
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        xlsx_path = os.path.join(out_dir, f"{base}_table.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "2theta-RelInt"

        # Header
        ws.append(["No.", "2θ (°)", "Rel. Intensity (%)"])
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for col in range(1, 4):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Data
        even_fill = PatternFill("solid", fgColor="D9E1F2")
        for i, (a, r) in enumerate(zip(angles, rel_intensity)):
            ws.append([i + 1, round(a, 4), round(r, 2)])
            if i % 2 == 1:
                for col in range(1, 4):
                    ws.cell(row=i + 2, column=col).fill = even_fill

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 20
        wb.save(xlsx_path)
        print(f"  [OK] XLSX: {xlsx_path}")
    except ImportError:
        print(f"  [--] XLSX: skipped (openpyxl not installed)")
        print(f"       Run: pip install openpyxl to enable XLSX output")

    print()
    print("=" * 60)
    print("  Done! 输出文件在同目录下。")
    print("=" * 60)
    input("Press Enter to exit...")


if __name__ == '__main__':
    main()
